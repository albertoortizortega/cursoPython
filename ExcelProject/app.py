import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        sheet.cell(row, 4).value = cell.value * 0.9
        sheet.cell(row, 4).number_format = '$#,##0.00'

    sheet.cell(1, 4).value = 'Price after discount'

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    wb.save(filename)

process_workbook('transactions.xlsx')

print("Workbook processed successfully.")